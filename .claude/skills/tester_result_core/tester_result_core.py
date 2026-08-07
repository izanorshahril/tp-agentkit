from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class ResultTable:
    source_format: str
    headers: list[str]
    data_rows: list[list[str]]
    id_row: list[str] | None = None

    def column_for_test_id(self, target_id: int) -> int:
        search_headers = self.id_row if self.id_row is not None else self.headers
        return resolve_column_index(search_headers, target_id)

    def column_index(self, column_name: str) -> int:
        return self.headers.index(column_name)

    def column_name(self, column_index: int) -> str:
        return self.headers[column_index].strip()

    def iter_rows(self):
        return iter(self.data_rows)

    def iter_rows_with_values(self, *column_indexes: int):
        highest_index = max(column_indexes, default=-1)
        for row in self.data_rows:
            if len(row) <= highest_index:
                continue
            if any(not row[column_index].strip() for column_index in column_indexes):
                continue
            yield row


def parse_active_ids(header_path: Path, symbol: str) -> list[int]:
    text = header_path.read_text(encoding="utf-8", errors="ignore")
    text = re.sub(r"//.*", "", text)
    match = re.search(rf"const\s+int\s+{re.escape(symbol)}\[.*?\]\s*=\s*\{{(.*?)\}};", text, re.S)
    if not match:
        raise ValueError(f"Could not find {symbol} in {header_path}")
    return [int(token) for token in re.findall(r"\d+", match.group(1))]


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_float(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None
    return float(text)


def parse_measurement_value(value: str) -> float | None:
    text = value.strip().replace(",", "")
    if not text:
        return None
    match = re.match(r"^([-+]?\d+(?:\.\d+)?)", text)
    if not match:
        return None
    return float(match.group(1))


def load_csv_blocks(csv_path: Path) -> tuple[list[str], list[str], list[list[str]]]:
    lines = csv_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    start = next(index for index, line in enumerate(lines) if line.startswith("Parameter,"))
    rows = list(csv.reader(lines[start:]))
    normalized = [[cell.strip() for cell in row] for row in rows]
    return normalized[0], normalized[1], normalized[6:]


def load_stdf_result_table(csv_path: Path) -> ResultTable:
    headers, id_row, data_rows = load_csv_blocks(csv_path)
    return ResultTable(source_format="stdf_csv", headers=headers, data_rows=data_rows, id_row=id_row)


def load_table_csv(csv_path: Path) -> list[list[str]]:
    text = csv_path.read_text(encoding="utf-8", errors="ignore")
    rows = list(csv.reader(text.splitlines()))
    return [[cell.strip() for cell in row] for row in rows]


def load_table_xlsx(xlsx_path: Path, sheet_name: str | None) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("XLSX table loading requires openpyxl to be installed.") from exc

    workbook = load_workbook(xlsx_path, data_only=False, read_only=True)
    try:
        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
            if worksheet is None:
                raise ValueError(f"No active worksheet found in {xlsx_path}")

        rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([normalize_cell(cell) for cell in row])
        return rows
    finally:
        workbook.close()


def load_tester_datalog_txt(txt_path: Path) -> list[list[str]]:
    lines = txt_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    sample_rows: list[dict[str, str]] = []
    current_sample: dict[str, str] | None = None

    for raw_line in lines:
        line = raw_line.rstrip()
        if line.startswith("***** Test Plan "):
            if current_sample:
                sample_rows.append(current_sample)
            current_sample = {}
            continue

        if current_sample is None:
            continue

        if line.startswith("Test No") or line.startswith("DEBUG:") or not line.strip():
            continue

        parts = re.split(r"\s{2,}", line.strip())
        if len(parts) < 4:
            continue

        test_number = parts[0].strip()
        if not test_number.isdigit():
            continue

        value = parse_measurement_value(parts[3])
        if value is None:
            continue
        current_sample[test_number] = str(value)

    if current_sample:
        sample_rows.append(current_sample)

    if not sample_rows:
        raise ValueError(f"No sample blocks parsed from {txt_path}")

    ordered_ids: list[str] = []
    seen: set[str] = set()
    for sample in sample_rows:
        for test_id in sample:
            if test_id not in seen:
                seen.add(test_id)
                ordered_ids.append(test_id)

    rows: list[list[str]] = [ordered_ids]
    for sample in sample_rows:
        rows.append([sample.get(test_id, "") for test_id in ordered_ids])
    return rows


def load_result_table(
    path: Path,
    kind: str,
    *,
    id_row: int = 1,
    sheet_name: str | None = None,
) -> ResultTable:
    if kind == "stdf_csv":
        return load_stdf_result_table(path)

    if kind == "csv":
        rows = load_table_csv(path)
    elif kind == "xlsx":
        rows = load_table_xlsx(path, sheet_name)
    elif kind == "tester_txt":
        rows = load_tester_datalog_txt(path)
    else:
        raise ValueError(f"Unsupported result-table kind: {kind}")

    if id_row < 1:
        raise ValueError("id_row must be >= 1")
    if len(rows) < id_row:
        raise ValueError("Input does not contain the requested ID row")

    id_index = id_row - 1
    headers = [cell.strip() for cell in rows[id_index]]
    data_rows = rows[id_index + 1 :]
    return ResultTable(source_format=kind, headers=headers, data_rows=data_rows, id_row=headers)


def resolve_column_index(headers: list[str], target_id: int) -> int:
    target = str(target_id)
    normalized_headers = [header.strip() for header in headers]
    for index, header in enumerate(normalized_headers):
        if header == target:
            return index

    token_pattern = re.compile(rf"(?<!\d){re.escape(target)}(?!\d)")
    matches = [index for index, header in enumerate(normalized_headers) if token_pattern.search(header)]
    if len(matches) == 1:
        return matches[0]
    raise ValueError(f"Could not uniquely resolve column for ID {target}")


__all__ = [
    "ResultTable",
    "load_csv_blocks",
    "load_result_table",
    "load_stdf_result_table",
    "load_table_csv",
    "load_table_xlsx",
    "load_tester_datalog_txt",
    "normalize_cell",
    "parse_active_ids",
    "parse_float",
    "parse_measurement_value",
    "resolve_column_index",
]

