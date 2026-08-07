from __future__ import annotations

import importlib.util
import sys
import tempfile
import textwrap
import unittest
from pathlib import Path


SKILLS_ROOT_PATH = Path(__file__).resolve().parents[1]
if str(SKILLS_ROOT_PATH) not in sys.path:
    sys.path.insert(0, str(SKILLS_ROOT_PATH))

from _test_support import SKILLS_ROOT, SkillTestCase


OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None


class TesterResultCoreSkillTests(SkillTestCase):
    def test_load_result_table_from_csv(self) -> None:
        module = self.load_module(
            "tester_result_core_under_test",
            SKILLS_ROOT / "tester_result_core" / "tester_result_core.py",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "results.csv"
            self.write_text(csv_path, "SITE,1001,1002\n0,1.0,2.0\n1,1.1,2.1\n")

            table = module.load_result_table(csv_path, "csv")

            self.assertEqual(table.source_format, "csv")
            self.assertEqual(table.column_for_test_id(1002), 2)
            self.assertEqual(
                list(table.iter_rows_with_values(0, 1, 2)),
                [["0", "1.0", "2.0"], ["1", "1.1", "2.1"]],
            )

    def test_parse_active_ids_from_header_symbol(self) -> None:
        module = self.load_module(
            "tester_result_core_symbols",
            SKILLS_ROOT / "tester_result_core" / "tester_result_core.py",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            header_path = Path(temp_dir) / "RelativePairs.h"
            self.write_text(
                header_path,
                textwrap.dedent(
                    """\
                    // Comment line should be ignored.
                    const int RelativeTestIDList[] = {1001, 1002, 1003};
                    """
                ),
            )

            self.assertEqual(module.parse_active_ids(header_path, "RelativeTestIDList"), [1001, 1002, 1003])

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
    def test_load_result_table_from_xlsx(self) -> None:
        module = self.load_module(
            "tester_result_core_xlsx",
            SKILLS_ROOT / "tester_result_core" / "tester_result_core.py",
        )

        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = Path(temp_dir) / "results.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            assert worksheet is not None
            worksheet.title = "Results"
            worksheet.append(["SITE", "1001", "1002"])
            worksheet.append(["0", "1.0", "2.0"])
            worksheet.append(["1", "1.1", "2.1"])
            workbook.save(xlsx_path)

            table = module.load_result_table(xlsx_path, "xlsx")

            self.assertEqual(table.source_format, "xlsx")
            self.assertEqual(table.column_for_test_id(1002), 2)
            self.assertEqual(
                list(table.iter_rows_with_values(0, 1, 2)),
                [["0", "1.0", "2.0"], ["1", "1.1", "2.1"]],
            )


if __name__ == "__main__":
    unittest.main(argv=[sys.argv[0]])